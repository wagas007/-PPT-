from __future__ import annotations

from datetime import date
from html import escape
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
TODAY = date.today().isoformat()

SOURCES = {
    "城市車旅聯絡我們": "https://www.cityparking888.com/contact.php",
    "城市車旅認識我們": "https://www.cityparking888.com/about.php",
    "臺灣證券交易所_阜爾運通新上市公司介紹": "https://wwwc.twse.com.tw/market_insights/zh/detail/ff8080818eead609018f27ed636f00ad",
    "阜爾運通公開說明書": "https://www.pss-group.com/upload/20230807095513.pdf",
    "城市車旅初步開發回報": "【城市車旅 × 旺來瓦斯｜初步開發回報】.docx",
    "既有案場資料庫": "CITY_案場資料庫_v1.xlsx",
    "供電評估模型": "停車場供電評估模型_修正版_v1.md",
}


def versioned_path(filename: str) -> Path:
    path = ROOT / filename
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    i = 2
    while True:
        candidate = ROOT / f"{stem}_v{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def doc_p(text: str, style: str | None = None) -> str:
    style_xml = f'<w:pStyle w:val="{style}"/>' if style else ""
    runs = []
    for idx, line in enumerate(str(text).split("\n")):
        if idx:
            runs.append("<w:r><w:br/></w:r>")
        runs.append(f'<w:r><w:t xml:space="preserve">{escape(line)}</w:t></w:r>')
    return f"<w:p><w:pPr>{style_xml}</w:pPr>{''.join(runs)}</w:p>"


def make_docx(path: Path, title: str, sections: list[tuple[str, list[str]]]) -> None:
    body = [
        doc_p(title, "Title"),
        doc_p(f"版本：v1｜產出日期：{TODAY}｜用途：城市車旅總部提案前DD與BD執行文件", "Subtitle"),
    ]
    for heading, paragraphs in sections:
        body.append(doc_p(heading, "Heading1"))
        for paragraph in paragraphs:
            body.append(doc_p(paragraph))
    doc_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>{''.join(body)}<w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="1134" w:right="1134" w:bottom="1134" w:left="1134"/></w:sectPr></w:body>
</w:document>"""
    styles = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:style w:type="paragraph" w:default="1" w:styleId="Normal"><w:name w:val="Normal"/><w:rPr><w:rFonts w:ascii="Arial" w:hAnsi="Arial" w:eastAsia="Microsoft JhengHei"/><w:sz w:val="21"/></w:rPr><w:pPr><w:spacing w:after="120" w:line="300" w:lineRule="auto"/></w:pPr></w:style>
  <w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/><w:rPr><w:rFonts w:eastAsia="Microsoft JhengHei"/><w:b/><w:sz w:val="32"/></w:rPr><w:pPr><w:spacing w:after="240"/></w:pPr></w:style>
  <w:style w:type="paragraph" w:styleId="Subtitle"><w:name w:val="Subtitle"/><w:rPr><w:rFonts w:eastAsia="Microsoft JhengHei"/><w:color w:val="666666"/><w:sz w:val="19"/></w:rPr></w:style>
  <w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/><w:basedOn w:val="Normal"/><w:rPr><w:rFonts w:eastAsia="Microsoft JhengHei"/><w:b/><w:sz w:val="26"/><w:color w:val="1F4E79"/></w:rPr><w:pPr><w:spacing w:before="260" w:after="120"/></w:pPr></w:style>
</w:styles>"""
    with ZipFile(path, "w", ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/><Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/></Types>""")
        zf.writestr("_rels/.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/></Relationships>""")
        zf.writestr("word/_rels/document.xml.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>""")
        zf.writestr("word/document.xml", doc_xml)
        zf.writestr("word/styles.xml", styles)


def load_city_rows() -> tuple[list[list], list[list]]:
    wb = load_workbook(ROOT / "CITY_案場資料庫_v1.xlsx", data_only=True, read_only=True)
    raw = []
    ws = wb["01_原始資料"]
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw.append(list(row))
    top = []
    ws = wb["03_TOP10"]
    for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
        top.append(list(row))
    return raw, top


def fmt(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for col in range(1, ws.max_column + 1):
        width = 10
        for cell in ws[get_column_letter(col)]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)), 44))
        ws.column_dimensions[get_column_letter(col)].width = width + 2
    ws.auto_filter.ref = ws.dimensions


def make_excel(path: Path) -> None:
    raw, top = load_city_rows()
    wb = Workbook()

    ws = wb.active
    ws.title = "01_Company_Profile"
    ws.append(["項目", "內容", "BD含意", "資料來源/待驗證"])
    company_rows = [
        ["品牌", "城市車旅 CITY PARKING，阜爾運通/PSS Group 旗下停車場經營管理品牌", "可用總部/集團角度溝通，不只找單一案場", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["母公司", "阜爾運通股份有限公司，股票代號6914", "上市公司治理，提案需重視合規、責任分工與可量化", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["定位", "停車場自動化設備製造、安裝、軟體科技、維護保養、經營管理、行銷服務垂直整合", "城市車旅可理解智慧設備與場域服務，適合談低干擾智慧服務據點", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["規模", "2023年服務突破1000場，管理超過12萬格停車位；為全台大型停車場管理業者", "具備標準化擴點價值，但不可一開始要求20點", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["技術能力", "雲端整合、線上支付、數據分析、API串聯、自助繳費與車牌辨識等", "提案應接上智慧場域、數據回顧、營運效率，而非單純租地", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["異業合作訊號", "公開說明書提及擴大停車後服務市場，與理念一致之異業結盟，為車主提供更多停車後服務", "旺來可定位為停車後生活服務/社區服務節點", SOURCES["阜爾運通公開說明書"]],
        ["官方合作入口", "官網聯絡頁含停車場委託經營、異業結盟、廣告、旅遊業、場地租借、停車特約等欄位", "可用異業結盟/場地租借/停車特約路徑收案", SOURCES["城市車旅聯絡我們"]],
        ["聯絡資訊", "sales@pss-group.com；台北02-2246-1708；台中04-2301-9699；24H服務0800-208-333 / 02-2246-2098", "首波應同時寄Email與電話追蹤台中/台北窗口", SOURCES["城市車旅聯絡我們"]],
    ]
    for r in company_rows:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("02_Business_Model")
    ws.append(["Business Model Canvas構面", "城市車旅現況判斷", "對旺來合作含意", "待驗證"])
    for r in [
        ["Customer Segments", "地主、企業、政府、大樓管委會、商場/醫療/生活場域、停車用戶", "提案需同時顧及場地業主、城市車旅營運端與車主/社區用戶", "各案場簽約主體"],
        ["Value Proposition", "穩定收益、專業停車物業管理、自動化與智慧化停車服務", "旺來合作要增加坪效/場域價值，不造成停車收益損失", "合作收益或品牌價值偏好"],
        ["Channels", "全台場站、官網、客服、區域業務、現場設備/APP/支付服務", "智慧自取櫃可作為停車後生活服務延伸", "是否允許場站導流與標示"],
        ["Key Activities", "場站開發、設備建置、維運、收費管理、數據管理、客戶服務", "城市車旅能理解設備維運，但仍需旺來承擔本案設備與客服", "雙方維運界線"],
        ["Key Resources", "停車場通路、車辨/繳費系統、區域業務、維修站、資料分析能力", "可要求對方協助提供場站推薦與供電資料", "是否可提供候選場站清單"],
        ["Revenue Streams", "停車費、委託管理、租斷承攬、設備/系統相關收入、異業合作可能收益", "先不主打租金，主打試點價值與後續合作可能", "分潤/固定合作費偏好"],
        ["Cost Structure", "場地成本、人員/維運、設備、系統、客服、工程", "若增加現場成本會降低接受度，需強調旺來負責營運", "用電與施工成本分攤"],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("03_Decision_Map")
    ws.append(["層級", "推測角色/單位", "可能關注", "旺來應準備", "取得方式", "狀態"])
    for r in [
        ["L1", "官網表單/總機/客服", "案件分類與是否轉派正確窗口", "一句話定位、3個候選案場、簡報邀約", "sales@pss-group.com、台中/台北電話", "待啟動"],
        ["L2", "區域業務/分案場業務", "該案場是否有空間、權責、供電與收益可能", "TOP3案場摘要、現勘表、供電三層問題", "依既有電話回報，城市車旅偏分區/分案場管理", "已知路線"],
        ["L3", "後勤/跨部門評估單位", "是否影響營運、品牌、法規、施工、安全", "總部提案、責任分工、退出機制、風險表", "由業務窗口轉介", "待驗證"],
        ["L4", "工程/維運/場站管理", "電源、車道、施工、補貨短停、設備位置", "現勘照片、電箱距離、補貨動線、設備規格", "現勘階段", "待驗證"],
        ["L5", "主管/決策者", "是否值得試點、是否可複製、是否帶來品牌/收益價值", "1至3點試點框架、90天KPI、半年擴點條件", "簡報或內部提報", "待取得簡報"],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("04_Competitive_Analysis")
    ws.append(["品牌", "定位", "優勢", "弱點/限制", "對旺來切入啟示"])
    for r in [
        ["城市車旅", "大型智慧停車/垂直整合管理品牌", "規模大、技術/設備/維修整合、官方異業入口、台中電話窗口", "分區/分案場決策可能造成總部推進需具體案場", "用具體案場+總部試點框架切入"],
        ["Times/PARK24", "日系標準化停車品牌", "標準化強、場站密度高、品牌管理嚴謹", "法務/安全審查較嚴", "城市車旅相較更能接受案場驅動與台中區域開發"],
        ["嘟嘟房", "中興電工體系停車品牌", "官網委託合作入口明確", "客服分流，安全疑慮需管理", "城市車旅可用既有電話接觸紀錄加速"],
        ["俥亭/ViVi PARK/其他", "區域或特定通路停車品牌", "彈性可能較高", "標準化與擴點規模不一定足", "城市車旅若成案，示範價值較強"],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("05_Fit_Analysis")
    ws.append(["Fit面向", "城市車旅特徵", "旺來合作價值", "分數", "判斷"])
    for r in [
        ["通路Fit", "全台大型停車場品牌，台中場站清單已盤點", "可作為社區型服務節點網絡", 5, "高"],
        ["營運Fit", "自助繳費、車辨、雲端管理與維運能力成熟", "能理解設備化服務，但旺來仍需承擔日常營運", 4, "高"],
        ["商務Fit", "公開資料支持異業結盟與停車後服務市場", "本案符合場站服務延伸", 4, "高"],
        ["供電Fit", "設備化場站可能已有電源，但旁空地/平面場供電需現勘", "需把供電作為A/B/C前置門檻", 3, "中"],
        ["權責Fit", "分區/分案場管理，可能需業務與總部雙軌", "先案場、後總部，或案場資料輔助總部提案", 3, "中"],
        ["ESG/智慧場域Fit", "公開資料提及低碳智慧停車、異業整合與停車後服務", "可包裝成智慧生活服務與低干擾場域活化", 4, "高"],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("06_Risk_Register")
    ws.append(["風險類型", "風險描述", "可能影響", "機率", "衝擊", "等級", "緩解措施", "Owner", "狀態"])
    for r in [
        ["供電", "平面/旁空地不一定有合法可用電源", "無法列A或施工成本高", "高", "高", "紅", "供電三層判斷；供電<=2不得列A", "工程/BD", "待現勘"],
        ["動線", "高人流商圈可能人車交織，設備影響車道或停車格", "城市車旅拒絕設置", "中", "高", "紅", "只選邊角、不占格、不擋主要動線", "BD/營運", "待現勘"],
        ["權責", "案場可能由地主/管委會/委託方持有決策權", "審查延長或無法簽約", "中", "高", "紅", "首輪即問簽約主體與同意權", "BD/法務", "待驗證"],
        ["現場負擔", "對方擔心現場管理員需處理設備/客訴", "降低合作意願", "中", "中", "黃", "旺來承擔客服、補貨、維運、異常", "BD/客服", "可控"],
        ["品牌安全", "瓦斯服務可能引發安全疑慮", "被後勤或主管擋下", "中", "高", "紅", "使用智慧服務據點語言；補安全/保險/合規文件", "BD/法務", "待補文件"],
        ["商務", "收益分配或租金期待不明", "談判延誤", "中", "中", "黃", "先談試點和數據，不先承諾大量租金模型", "BD/主管", "待談"],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("07_News_Research")
    ws.append(["主題", "重點", "BD含意", "來源"])
    for r in [
        ["上市公司介紹", "阜爾運通於2024/05/03上市；以城市車旅從事停車場經營管理，完成垂直整合", "用上市公司/總部治理語氣提案", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["規模", "2023年服務突破1000場，管理超過12萬格停車位", "具備示範成功後擴點空間", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["技術", "雲端整合、線上支付、數據分析、API串聯", "可把旺來服務定位成智慧場域服務延伸", SOURCES["臺灣證券交易所_阜爾運通新上市公司介紹"]],
        ["異業整合", "公開說明書提及停車後服務與異業結盟，包括電動車充電、維修保養、汽車美容、保險等", "旺來可主張停車後/社區生活服務合作", SOURCES["阜爾運通公開說明書"]],
        ["官方合作欄位", "官網聯絡頁含異業結盟、場地租借、停車特約等欄位", "可用官方入口正式收案", SOURCES["城市車旅聯絡我們"]],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("08_Action_Plan")
    ws.append(["階段", "時間", "任務", "負責人", "輸入", "輸出", "成功標準", "狀態"])
    for r in [
        ["Phase 1", "D+1", "寄送城市車旅3案場提案Email", "BD", "HEADQUARTER_APPROACH/TOP10", "寄件紀錄", "對方回覆或分派窗口", "待辦"],
        ["Phase 1", "D+1", "電話追蹤台中04-2301-9699與台北02-2246-1708", "BD", "寄件內容", "通話紀錄", "取得區域/業務窗口", "待辦"],
        ["Phase 2", "D+2", "請窗口確認TOP3案場權責與供電資料", "BD/工程", "TOP3/供電表", "待驗證清單", "取得現勘可行點", "待辦"],
        ["Phase 2", "D+3", "安排1至3個案場現勘", "BD/營運/工程", "現勘表", "照片/供電/動線紀錄", "至少1案符合試點條件", "待辦"],
        ["Phase 3", "D+5", "邀約總部或後勤30分鐘簡報", "BD主管", "DD/BD/Approach文件", "會議邀約", "取得總部簡報時間", "待辦"],
        ["Phase 4", "D+7", "整理試點條件與分工表", "PMO/法務", "現勘結果", "試點框架", "進入合作條件討論", "待辦"],
    ]:
        ws.append(r)
    fmt(ws)

    ws = wb.create_sheet("09_TOP10_From_Existing")
    ws.append(["排序", "編號", "場站名稱", "地址", "推薦原因", "風險", "試行方式", "導入條件", "預估合作價值"])
    for r in top:
        ws.append(r[:9])
    fmt(ws)

    wb.save(path)


def source_text() -> str:
    return "\n".join([f"{k}: {v}" for k, v in SOURCES.items()])


def top10_text() -> str:
    _, top = load_city_rows()
    return "\n".join([f"{r[0]}. {r[2]}｜{r[3]}｜理由：{r[4]}｜風險：{r[5]}" for r in top[:10]])


def dd_sections() -> list[tuple[str, list[str]]]:
    return [
        ("0. Executive Summary", [
            "結論：城市車旅是目前最適合先推進到「可進行總部/區域提案」的品牌。原因有三：第一，既有電話回報顯示對方開放評估，但偏好以具體案場切入；第二，城市車旅/阜爾運通具備大型停車通路、設備化與智慧化能力；第三，公開資料明確支持異業結盟與停車後服務延伸。",
            "建議策略：採「案場驅動型總部提案」。先帶3個台中候選示範點，請城市車旅分派對應區域/案場業務，完成供電、權責與動線初審後，再邀約總部或後勤單位進行30分鐘正式簡報。",
            "第一波不要求20點。標準說法為：先以1至3個示範場站驗證；若空間、供電、需求與營運數據成立，未來半年再評估台中與桃園約20個合作據點。",
        ]),
        ("1. Company Profile", [
            "城市車旅 CITY PARKING 是阜爾運通/PSS Group旗下停車場經營管理品牌。臺灣證券交易所新上市公司介紹指出，阜爾運通以停車場自動化設備製造、銷售為基礎，並以城市車旅從事停車場經營管理服務，完成工廠製造、安裝服務、軟體科技、維護保養、經營管理、行銷服務的垂直整合。",
            "公開資料顯示，阜爾運通在2023年服務突破1000場、管理超過12萬格停車位。這代表城市車旅具備場站規模與標準化能力，若試點成立，後續具有複製價值。",
            "城市車旅官網聯絡頁列出異業結盟、場地租借、停車特約等合作欄位，並公開 sales@pss-group.com、台北與台中聯絡電話。這是本案第一個正式進件入口。",
        ]),
        ("2. Business Due Diligence", [
            "商業模式：城市車旅不是單純停車場品牌，而是停車場設備、系統、經營管理、維運與行銷整合商。旺來提案若只講租角落，會低估對方的智慧場域能力；應改以場站增值服務與停車後生活服務切入。",
            "合作價值：對城市車旅而言，旺來可提供不增加現場負擔的社區型智慧服務據點，活化不影響車格/車道的邊角空間，並提供可量化的試點數據。",
            "關鍵限制：城市車旅既有接觸紀錄顯示其採業務分區/分案場管理，因此推進方式需同時具備具體案場清單與總部標準化敘事。",
        ]),
        ("3. Fit Analysis", [
            "高Fit：通路規模、智慧設備能力、官方異業合作入口、停車後服務延伸方向。",
            "中Fit：分區/分案場決策使總部一次性框架較難立即成立；供電可行性仍需逐案驗證。",
            "低Fit風險：若案場需占車格、影響主要動線、需新增電表或地主/管委會權責不清，該案不應進入A級試點。",
        ]),
        ("4. Recommended First Wave", [
            "第一波建議從既有TOP10中挑3個作為對外提案材料：一中二站、惠中A站、吉拾忠明站/大潤發忠明站。若城市車旅業務認為區域權責不同，可請其改推薦台中3個室外、具邊角空間且有既有供電的候選案場。",
            "TOP10摘要：\n" + top10_text(),
        ]),
        ("5. Risk Register Summary", [
            "主要風險依序為供電、動線、權責、品牌安全、現場負擔與商務條件。供電風險必須前置處理：有空間不代表可設置；供電可行性<=2不得列A。",
            "城市車旅設備化程度高，許多場站可能已有車辨、繳費、照明或管理亭電源，但旺來不得假設可直接分接，需取得合法用電方式與費用結算方式。",
        ]),
        ("6. Sources", [source_text()]),
    ]


def bd_sections() -> list[tuple[str, list[str]]]:
    return [
        ("0. BD Strategy Conclusion", [
            "城市車旅BD策略應採「案場驅動型總部提案」：不是直接要求總部給20個點，而是先用台中3個候選案場建立具體討論，再請業務/區域窗口分派，最後進入總部或後勤正式評估。",
            "第一階段成功標準：取得城市車旅業務或區域窗口，並安排30分鐘簡報或1至3個案場現勘。",
        ]),
        ("1. Entry Route", [
            "入口一：寄送Email至 sales@pss-group.com，主旨建議為「旺來瓦斯 × 城市車旅｜台中示範場站智慧服務據點合作評估」。",
            "入口二：電話追蹤台中04-2301-9699，因本案首波案場在台中，需請轉台中區域業務/場站合作窗口。",
            "入口三：電話追蹤台北02-2246-1708，請總部協助確認異業結盟/場地租借/業務開發承辦。",
        ]),
        ("2. Account Plan", [
            "首波提案不寄完整DD長文，而是寄：合作摘要、TOP3案場、旺來負責事項、城市車旅協助事項、30分鐘簡報邀約。",
            "BD話術重點：我們不是要租停車格，也不是請現場管理員處理設備，而是希望先確認1至3個示範點是否具備邊角空間、合法供電與不影響動線的條件。",
        ]),
        ("3. TOP3 Recommendation", [
            "建議TOP3：一中二站、惠中A站、吉拾忠明站/大潤發忠明站。備選：精武東站、台中公園站、中港文心站。",
            "選點邏輯：生活密度高、商圈/住宅訊號明確、具平面或旁空地訊號、可作為示範點；但所有供電、權責、實際邊角空間均列待驗證。",
        ]),
        ("4. Conversion Path", [
            "D+1：寄信與電話追蹤。D+2：取得承辦窗口或補件要求。D+3：確認TOP3案場權責與供電資訊。D+5：安排現勘或30分鐘簡報。D+7：形成1至3點試點條件表。",
            "若總部推進慢：改走案場/區域路線補資料，但不得在未取得權責確認前承諾設置或條件。",
        ]),
        ("5. KPI", [
            "BD KPI：取得窗口、取得簡報、取得現勘、取得供電資料、確認1至3個試點候選。",
            "試點KPI：取貨/使用量、補貨效率、異常率、客訴率、場站回饋、供電穩定性、是否影響停車營運。",
        ]),
        ("6. Sources", [source_text()]),
    ]


def approach_sections() -> list[tuple[str, list[str]]]:
    return [
        ("1. 對城市車旅的提案主張", [
            "旺來瓦斯希望與城市車旅評估建立「社區型智慧服務據點」試點合作。此合作不是承租停車格，也不是要求大量點位，而是在不影響停車收益、車道動線與現場管理的前提下，利用場站閒置邊角空間導入智慧自取服務。",
            "旺來負責設備、客服、補貨、維運與異常處理；城市車旅協助確認候選場站、場地權責、合法供電與現勘窗口。",
        ]),
        ("2. 為什麼適合城市車旅", [
            "城市車旅具備大型場站通路、智慧停車設備、雲端與維運能力，且公開資料支持異業結盟與停車後服務延伸。本案可作為低干擾、可量化、可複製的生活服務延伸。",
            "對城市車旅而言，合作價值不是單一租金，而是提升場站服務價值、活化低使用角落、增加智慧場域形象，並維持停車本業不受影響。",
        ]),
        ("3. 試點方式", [
            "建議先以1至3個示範點試行。候選條件：室外或地面層、具邊角空間、不影響車格/車道、靠近住宅或商圈、已有合法可協調電源、補貨可短停、權責清楚。",
            "試點期間建議90天，追蹤KPI：使用量、補貨效率、異常率、客訴率、場站回饋、用電穩定性、是否影響停車營運。",
        ]),
        ("4. 首波候選案場", [
            "初步建議以一中二站、惠中A站、吉拾忠明站/大潤發忠明站作為討論素材。若城市車旅有更適合的台中室外場站，建議由業務窗口直接提供3至5個候選點，旺來依同一套標準評估。",
            "所有候選場站均需確認供電、權責、動線與補貨條件，未現勘前不承諾設置。",
        ]),
        ("5. 雙方分工", [
            "旺來：設備、商品/服務營運、客服、補貨、維運、異常處理、保險與安全文件、試點數據回顧。",
            "城市車旅：候選場站推薦、場地權責確認、供電可行性協助、現勘窗口、品牌/後勤/工程審查流程確認。",
        ]),
        ("6. 下一步", [
            "請城市車旅協助安排30分鐘總部/區域簡報。會議目標不是立即簽約，而是確認：承辦窗口、候選場站、供電資料、現勘流程、試點KPI、責任分工與下一次評估會議。",
            "若試點成立，未來半年可依標準化評估表，逐步評估台中與桃園約20個合作據點；不建議在試點前承諾大量點位。",
        ]),
        ("7. 資料來源", [source_text()]),
    ]


def main() -> None:
    outputs = [
        (versioned_path("CITY_PARKING_DD_MASTER.xlsx"), make_excel),
        (versioned_path("CITY_PARKING_DD_REPORT.docx"), lambda path: make_docx(path, "CITY PARKING Due Diligence Report", dd_sections())),
        (versioned_path("CITY_PARKING_BD_STRATEGY.docx"), lambda path: make_docx(path, "CITY PARKING BD Strategy", bd_sections())),
        (versioned_path("CITY_PARKING_HEADQUARTER_APPROACH.docx"), lambda path: make_docx(path, "CITY PARKING Headquarter Approach", approach_sections())),
    ]
    for path, maker in outputs:
        maker(path)
        print(path.name)


if __name__ == "__main__":
    main()
