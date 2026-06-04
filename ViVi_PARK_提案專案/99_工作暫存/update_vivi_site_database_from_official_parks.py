# -*- coding: utf-8 -*-
"""Update ViVi PARK site database from the official /parks page HTML.

Input: a UTF-8 HTML file saved from https://vivi-park.com/parks/
Output: ViVi_PARK_提案專案/00_來源資料/VIVI_案場資料庫_v1.xlsx
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "00_來源資料" / "VIVI_案場資料庫_v1.xlsx"


def load_parks(html_path: Path) -> list[dict]:
    html = html_path.read_text(encoding="utf-8")
    match = re.search(r"let\s+allParks\s*=\s*(\[.*?\]);", html, re.S)
    if not match:
        raise RuntimeError("找不到 let allParks = [...]；官方頁面結構可能已改版。")
    return json.loads(match.group(1))


def infer_site_type(name: str, address: str) -> str:
    text = f"{name}{address}"
    if any(k in text for k in ["B1", "B2", "B3", "地下"]):
        return "地下／室內"
    if any(k in text for k in ["全聯", "家樂福", "愛買", "商場"]):
        return "零售／賣場場站"
    if any(k in text for k in ["空地", "旁", "對面", "路口", "交叉"]):
        return "室外平面／空地型"
    return "室外或一般場站（待現勘）"


def score_park(park: dict) -> dict:
    name = park.get("pa_name") or ""
    county = park.get("county") or ""
    city = park.get("city") or ""
    address = park.get("address") or ""
    full = f"{county}{city}{address}"
    text = f"{name}{full}"
    month_fee = int(park.get("month_fee") or 0)

    open_signal = any(k in text for k in ["空地", "旁", "對面", "路口", "交叉", "前"])
    underground = any(k in text for k in ["B1", "B2", "B3", "地下", "室內"])
    retail = any(k in name for k in ["全聯", "家樂福", "愛買"])
    transit = any(k in text for k in ["高鐵", "車站", "火車"])
    residential = any(k in city for k in ["桃園", "中壢", "平鎮", "八德", "蘆竹", "北屯", "南屯", "潭子", "大里", "豐原", "烏日"])
    urban_core = any(k in text for k in ["同安", "藝文", "梅川", "北屯", "潭子", "大里", "中壢", "平鎮", "大墩", "惠中", "松義"])

    value = 2
    if residential:
        value += 1
    if urban_core:
        value += 1
    if retail:
        value += 1
    if month_fee > 0:
        value += 0.5
    if transit and not retail:
        value -= 0.5
    value = max(1, min(5, round(value)))

    if underground:
        difficulty = 5
    elif open_signal and not retail:
        difficulty = 1
    elif retail:
        difficulty = 3
    else:
        difficulty = 3

    # Strategic fit adds priority nuance without hiding operational risk.
    strategic = 0
    if month_fee > 0:
        strategic += 1
    if retail:
        strategic += 1
    if open_signal:
        strategic += 1

    total = value * 20 + (6 - difficulty) * 10 + strategic * 5
    if value >= 4 and difficulty <= 1:
        priority = "A"
    elif value >= 4 and difficulty <= 3:
        priority = "B+"
    elif value >= 3:
        priority = "B"
    else:
        priority = "C"

    if underground:
        open_space = "否"
        edge = "低【地下／室內】"
    elif open_signal:
        open_space = "是"
        edge = "高【地址具空地／旁／路口訊號】"
    elif retail:
        open_space = "待現勘"
        edge = "中【零售場權責待確認】"
    else:
        open_space = "待現勘"
        edge = "中【待現勘】"

    risk = []
    if retail:
        risk.append("零售／第三方權責")
    if underground:
        risk.append("地下或室內導入阻力")
    if transit:
        risk.append("通勤型需求波動")
    if not risk:
        risk.append("供電與動線待現勘")

    return {
        "site_type": infer_site_type(name, full),
        "open_space": open_space,
        "edge": edge,
        "residential": "是" if residential else "待驗證",
        "commerce": "是" if (retail or urban_core) else "待驗證",
        "market": "待驗證",
        "school": "待驗證",
        "power": "高【官方場站具停車設備，仍需現勘分接】" if not underground else "中【地下／室內需確認】",
        "value": value,
        "difficulty": difficulty,
        "priority": priority,
        "total": total,
        "risk": "、".join(risk),
        "next": "請 ViVi PARK 提供電箱位置、場站權責與可現勘窗口",
    }


def rows_for_city(parks: list[dict], county: str, prefix: str) -> list[list]:
    filtered = [p for p in parks if p.get("county") == county]
    rows = []
    for idx, park in enumerate(filtered, 1):
        s = score_park(park)
        name = park.get("pa_name") or ""
        city = park.get("city") or ""
        address = park.get("address") or ""
        full_addr = f"{park.get('county','')}{city}{address}"
        rows.append([
            f"{prefix}-{idx:03d}",
            park.get("pa_id"),
            name,
            full_addr,
            city,
            s["site_type"],
            park.get("month_fee") or 0,
            s["open_space"],
            s["edge"],
            s["residential"],
            s["commerce"],
            s["market"],
            s["school"],
            s["power"],
            s["value"],
            s["difficulty"],
            s["priority"],
            s["total"],
            park.get("pa_latitude"),
            park.get("pa_Longitude"),
            s["risk"],
            s["next"],
            "官方 /parks allParks",
        ])
    rows.sort(key=lambda r: (-r[17], r[15], r[2]))
    return rows


def style(ws, widths: list[int]):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        if len(row) > 16:
            pri = row[16].value
            fill = None
            if pri == "A":
                fill = "C6EFCE"
            elif pri == "B+":
                fill = "DDEBF7"
            elif pri == "B":
                fill = "FFF2CC"
            elif pri == "C":
                fill = "FCE4D6"
            if fill:
                row[16].fill = PatternFill("solid", fgColor=fill)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_sheet(ws, headers: list[str], rows: list[list], widths: list[int]):
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style(ws, widths)


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python update_vivi_site_database_from_official_parks.py <parks_html>")

    html_path = Path(sys.argv[1])
    parks = load_parks(html_path)
    taichung = rows_for_city(parks, "台中市", "VIVI-TC")
    taoyuan = rows_for_city(parks, "桃園市", "VIVI-TY")

    wb = load_workbook(XLSX)
    headers = [
        "編號", "官方ID", "場站名稱", "完整地址", "區域", "場站型態推定", "官方月租費",
        "是否開放空間", "邊角/空地訊號", "周邊住宅", "商圈/零售", "市場", "學區",
        "供電訊號", "合作價值(1-5)", "導入難度(1/3/5)", "優先級", "總分",
        "緯度", "經度", "主要風險", "下一步", "資料來源",
    ]
    widths = [13, 10, 24, 42, 14, 20, 12, 14, 24, 12, 14, 10, 10, 30, 14, 16, 10, 10, 13, 13, 28, 42, 20]

    write_sheet(wb["03_台中場站庫"], headers, taichung, widths)
    write_sheet(wb["04_桃園場站庫"], headers, taoyuan, widths)

    top = sorted(taichung + taoyuan, key=lambda r: (-r[17], r[15], r[2]))[:10]
    top_rows = []
    for rank, r in enumerate(top, 1):
        recommendation = (
            "首波示範點候選" if r[16] in ("A", "B+") else "備選，需補現勘資料"
        )
        top_rows.append([
            rank, r[0], "台中" if r[0].startswith("VIVI-TC") else "桃園",
            r[2], r[3], r[5], r[6], r[14], r[15], r[16], r[17],
            f"{r[9]}／{r[10]}／{r[8]}", r[20], r[21], recommendation,
        ])
    write_sheet(
        wb["06_TOP10初選"],
        ["排序", "編號", "城市", "場站名稱", "地址", "場站型態", "月租費", "合作價值", "導入難度", "優先級", "總分", "推薦原因", "主要風險", "現勘必查", "提案語"],
        top_rows,
        [8, 13, 10, 24, 42, 20, 10, 10, 10, 10, 10, 40, 30, 42, 28],
    )

    power_rows = []
    for r in top:
        availability = "官方場站具停車設備，推測有基礎電源；仍不得假設可直接分接"
        accessibility = "若設備位置距電箱 <5M 且不跨車道，優先可行；零售場需確認第三方權責"
        economics = "短距離分接可控；需新增電表、跨車道拉線或商場審核則降級"
        score = 4 if r[16] == "A" else 3 if r[16] == "B+" else 2
        power_rows.append([r[0], r[2], availability, accessibility, economics, score, r[20], r[21]])
    power_rows.append([
        "所有場站", "通用判斷邏輯",
        "有車辨、繳費、照明或 APP 場站只代表可能有電，不代表可合法接電",
        "需確認電箱距離、分接合法性、線路是否跨車道與是否需第三方同意",
        "設備用電可控；但新設電表、長距離配線與第三方審核會拉高成本",
        "依現勘", "三層判斷：Availability → Accessibility → Economics",
        "低可行性不列入 1–3 試點；高可行性優先安排現勘",
    ])
    write_sheet(
        wb["07_供電分析"],
        ["編號", "場站名稱", "Availability 有沒有電", "Accessibility 能不能接", "Economics 是否划算", "供電分數", "判斷", "下一步"],
        power_rows,
        [13, 24, 42, 44, 42, 12, 32, 42],
    )

    source_rows = [
        ["官方 /parks allParks", "https://vivi-park.com/parks/", "桃園／台中官方場站名稱、地址、座標、月租費", "高", f"本次解析官方內嵌 allParks，共 {len(parks)} 筆；台中 {len(taichung)} 筆、桃園 {len(taoyuan)} 筆"],
        ["評分模型", "旺來 BD 場站評估模型", "合作價值、導入難度、優先級、供電初評", "內部假設", "需以 ViVi PARK 權責、供電與現勘結果確認"],
    ]
    write_sheet(
        wb["10_Source_Assumptions"],
        ["來源", "網址／依據", "用於", "可信度", "備註"],
        source_rows,
        [24, 46, 42, 12, 70],
    )

    # Update action plan status if present.
    if "09_Action_Plan" in wb.sheetnames:
        ws = wb["09_Action_Plan"]
        for row in ws.iter_rows(min_row=2):
            task = str(row[2].value or "")
            if "場站" in task or "案場" in task:
                row[7].value = "已依官方 /parks allParks 更新；待現勘確認"

    wb.save(XLSX)
    print(f"updated: {XLSX}")
    print(f"official parks: {len(parks)} / 台中 {len(taichung)} / 桃園 {len(taoyuan)}")


if __name__ == "__main__":
    main()
