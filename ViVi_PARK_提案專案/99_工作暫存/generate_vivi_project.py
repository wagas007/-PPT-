# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "00_來源資料"
ONE = ROOT / "01_OnePager"
COLD = ROOT / "02_Coldmail"
MEET = ROOT / "03_取得窗口與會議"
SHORT = ROOT / "04_簡版PPT"
FORMAL = ROOT / "05_正式PPT"
TMP = ROOT / "99_工作暫存"

for folder in [SRC, ONE, COLD, MEET, SHORT, FORMAL, TMP]:
    folder.mkdir(parents=True, exist_ok=True)

SOURCES = """## 主要公開來源

- ViVi PARK 官網首頁：智慧停車、路停、加油、EV 充電、生活服務、Vi 幣與聯絡資訊。https://vivi-park.com/
- ViVi PARK 關於我們：榮帝科技、董事長陳明賢、場站規模、ViVi 共榮生態圈、里程碑。https://vivi-park.com/about/
- ViVi PARK 同業聯盟：土地出租、同業加盟、地主加盟、非地主加盟。https://vivi-park.com/alliance-park/
- ViVi PARK 生活聯盟：鄰近停車場店家可在 APP 銷售商品券或點數支付。https://vivi-park.com/alliance-life/
- ViVi PARK 點數漫遊：Hami Point、momo 幣、P 幣、Super Point、亞洲萬里通里數等點數兌換 Vi 幣。https://vivi-park.com/alliance-point/
- ViVi PARK 停車場查詢頁：城市篩選與場站查詢入口。https://vivi-park.com/parks/
- 官方新場資訊：同安街場，桃園市桃園區同安街322號。https://vivi-park.com/2025/11/%E3%80%90%E6%96%B0%E5%A0%B4%E9%96%8B%E5%B9%95%E3%80%91vivi-park%E5%90%8C%E5%AE%89%E8%A1%97%E5%81%9C%E8%BB%8A%E5%A0%B4/
- 官方新場資訊：全聯平鎮中豐店，桃園市平鎮區中豐路南勢二段326號。https://www.vivi-park.com/News_Detail.aspx?News_ID=50
- 官方新場資訊：中壢松義二街，桃園市中壢區松義二街32號旁。https://old.vivi-park.com/News_Detail.aspx?News_ID=205
- 官方新場資訊：全聯台中潭子福潭店，台中市潭子區福潭路362、368號。https://www.vivi-park.com/News_Detail.aspx?News_ID=97
- 官方新場資訊：全聯台中大里爽文店，台中市大里區爽文路879號。https://vivi-park.com/News_Detail.aspx?News_ID=99
- 公司登記公開資料：榮帝科技股份有限公司統編 24963309、地址、董事長。https://data.bznk.com/24963309
"""

FILES = {}

FILES[SRC / "VIVI_DD_EXECUTION_GUIDE.md"] = f"""# ViVi PARK 合作開發 DD 執行指南

## Project Objective

本專案目標不是取得單一停車場空間，而是建立旺來瓦斯與 ViVi PARK／榮帝科技的總部合作模式。

核心假設：

```text
ViVi PARK = 停車場網絡 + Vi 幣點數 + 生活服務平台
旺來 = 可落地的社區生活服務供給與營運能力
合作主軸 = 把旺來服務接入 ViVi PARK 的場站與生活服務生態圈
```

## DD Scope

本次採商業合作 DD，重點為：

1. 公司與品牌定位
2. 場站與聯盟模式
3. Vi 幣、生活服務、點數漫遊與會員生態
4. 桃園／台中可試點案場方向
5. 總部合作切入路徑

## 禁止說法

- 租停車格
- 廣告進駐
- 傳統設備投放
- 一開始大量展點

## 正確說法

- 生活服務節點
- ViVi 共榮生態圈的場站服務延伸
- 1 至 3 個示範場站，90 天驗證
- 雙方共同決定是否擴大、調整或停止

{SOURCES}
"""

FILES[SRC / "VIVI_DD_REPORT.md"] = f"""# ViVi PARK 公開情報 DD 報告

## 1. 公司基本資料

| 項目 | 內容 | BD 含意 |
|---|---|---|
| 品牌 | ViVi PARK | 全國性智慧停車與生活服務平台 |
| 法人 | 榮帝科技股份有限公司 | 應以總部合作、平台合作角度切入 |
| 統編 | 24963309（公開公司登記資料） | 可作正式開發對象 |
| 董事長 | 陳明賢 Paul Chen | 官網明確揭露，品牌敘事集中 |
| 成立 | 2015 年 | 約十年品牌，已進入規模化與平台化階段 |
| 規模 | 官網新版稱自營逾 300 場、聯盟 700 場，小計超過 70 萬格；官網首頁亦稱串接超過 160 場、支援 15 縣市路邊停車 | 數字版本需以會議中官方口徑確認，BD 文案可用「全台智慧停車網絡」避免過度精確 |
| 聯絡 | 02-2536-6655；service-vivipark@kpclc.com.tw | 可作 cold mail 與電話追蹤入口 |

## 2. 商業模式分析

ViVi PARK 的收入與策略不只來自停車場管理，而是停車場、APP 會員、Vi 幣點數、特約商店、加油、EV 充電與同業聯盟的生態系。

| 模組 | 公開資訊 | 對旺來含意 |
|---|---|---|
| 停車場營運 | 自營、聯盟、地主加盟、同業加盟 | 可談總部框架，不只單點 |
| APP／Vi 幣 | 停車、路停、加油、充電、生活消費皆可使用 | 旺來可成為生活服務使用場景 |
| 生活服務 | 官網歡迎鄰近停車場店家以商品券或點數支付形式上架 | 旺來可從「商品券／服務節點」雙路徑切入 |
| 點數漫遊 | Hami Point、momo 幣、P 幣、Super Point、亞洲萬里通里數等 | 可把旺來服務包成點數消費場景 |
| EV／加油 | Vi POWER、合作加油與能源服務 | 旺來能源背景可與「車後生活服務」互補 |

## 3. 場站開發模式

ViVi PARK 明確提供土地出租、同業加盟、地主加盟、非地主加盟與管理系統導入。這代表其合作語言偏「聯盟、導流、平台、加盟」，不是單純租金模式。

## 4. 市場地位與差異化

| 品牌 | 差異化 | 旺來切入角 |
|---|---|---|
| PARK24 / Times | 國際品牌與標準化停車管理 | 總部試點與場站價值 |
| 城市車旅 | 大型場站營運與標準化管理 | 場站低干擾服務 |
| 嘟嘟房 | 品牌認知與 Pre-IPO 敘事 | 場站加值案例 |
| ViVi PARK | 停車＋會員點數＋生活服務平台 | 旺來服務接入 ViVi 共榮生態圈 |

## 5. 合作適配性

### Green Flags

- ViVi PARK 已公開主張生活服務與點數生態。
- 官網明確歡迎生活商家加入 APP 銷售商品券或點數支付。
- 已有 EV、加油、特約商店、點數漫遊合作經驗。
- 場站多為 APP 與車辨系統連動，具數位化基礎。

### Red Flags

- 平台合作可能有上架、點數清算與分潤規則。
- 若從場站放置設備切入，可能落入場地權責、地主同意與設置審核。
- 全聯、商場或聯盟場站權責較複雜。
- ViVi PARK 已有生活服務生態，旺來需清楚說明差異價值。

## 6. 總部合作可能性

值得進行總部接觸。建議開發入口分兩條線：

1. **生活服務／特約商店線**：旺來服務可作為 Vi 幣可消費或商品券合作。
2. **場站試點線**：1 至 3 個桃園／台中場站驗證社區型智慧服務節點。

## 7. 最終評分

評分：**A-，強烈建議投入，但必須把合作語言改成平台型。**

理由：

- ViVi PARK 的公開定位天然支援旺來的生活服務主張。
- 相較其他停車品牌，ViVi PARK 更容易理解「場站不只是停車」。
- 首波切入不應只講設備，而應同時講「Vi 幣會員、生活服務、場站服務節點」。

{SOURCES}
"""

FILES[SRC / "VIVI_BD_STRATEGY.md"] = """# ViVi PARK BD 開發策略稿

## 一句話策略

把旺來瓦斯從「想進停車場的服務商」，包裝成 ViVi PARK 生活服務生態圈中可落地、可驗證、可帶來新增使用場景的合作夥伴。

## 合作定位

```text
ViVi PARK × 旺來瓦斯
停車後生活服務節點試點合作
```

## 為什麼找 ViVi PARK

1. ViVi PARK 已經不是單純停車品牌，而是 APP、Vi 幣、生活服務、點數漫遊平台。
2. 官網明確開放生活服務與特約商店加入，旺來具備明確場景。
3. 停車場本身有社區、通勤、採買與生活圈接觸點，適合做 90 天小規模驗證。
4. ViVi PARK 需要持續擴大 Vi 幣使用場景，旺來可提供日常剛性需求服務。

## 建議切入順序

1. Cold mail 至客服信箱與官網表單，主旨鎖定「生活服務節點」而非「租用場地」。
2. 附 OnePager DM，讓窗口理解本案同時具備 APP 服務與場站驗證價值。
3. 會議中先談合作模式，不先指定大量場站。
4. 取得 1 至 3 個桃園／台中場站現勘或資料確認。
5. 90 天後以數據回顧：使用量、客訴、補貨、動線、Vi 幣／會員整合可能。

## 不適合一開始談

- 全台展點。
- 要求 ViVi PARK 直接上架或直接分潤。
- 場地租金細節。
- SOP 與大量設備規格。
- 未確認權責前指定商場或全聯場站一定可做。
"""

FILES[SRC / "VIVI_PROJECT_MASTER.md"] = """# ViVi PARK Project Master

## 專案目標

取得 ViVi PARK／榮帝科技窗口，安排 30 分鐘會議，討論旺來服務接入 ViVi PARK 場站與生活服務生態圈的可行性。

## 決策鏈假設

| 層級 | 角色 | 可能關注 | 接觸方式 |
|---|---|---|---|
| L1 | 董事長／高階主管 | 生態圈、品牌策略、平台價值 | 不直接第一線接觸 |
| L2 | 生活服務／聯盟合作主管 | 商品券、點數、導流、會員價值 | 主要切入 |
| L3 | 停車場營運／場站開發主管 | 場站條件、權責、動線、供電 | 試點確認 |
| L4 | 客服／行政窗口 | 轉介與收件 | 官網與客服信箱 |
| L5 | 法務／財務 | 合約、保險、清算、責任 | 後期進入 |

## 初步候選案場策略

優先城市：桃園、台中。

優先型態：

- 室外平面。
- 住宅／商圈／全聯生活場域。
- 可使用 ViVi PARK APP 或車辨設備。
- 已公開有臨停／月租，表示使用者穩定。

## 主要風險

1. 平台合作與場站合作權責不同。
2. 全聯或商場型場站可能需第三方同意。
3. Vi 幣點數合作有清算與上架門檻。
4. 設備放置需供電、保險與安全確認。
"""

FILES[SRC / "VIVI_HEADQUARTER_APPROACH.md"] = """# ViVi PARK 總部接觸策略

## 建議入口

- 客服信箱：service-vivipark@kpclc.com.tw
- 電話：02-2536-6655
- 官網同業聯盟／生活合作頁面
- LINE@：@vivi_park（作為補充，不作正式主入口）

## 第一封信目標

不是要求合作，也不是要求設點，而是取得正確窗口與 30 分鐘初步會議。

## 開場語氣

應從 ViVi PARK 的平台定位切入：

```text
我們留意到 ViVi PARK 正在拓展停車、Vi 幣、點數漫遊與生活服務生態圈。
旺來希望交流一個可以先小規模驗證的合作方向：把社區生活服務接入 ViVi PARK 的場站與會員使用場景。
```

## 會議要拿到的資訊

1. 生活服務／特約商店合作窗口。
2. 桃園／台中可評估場站清單。
3. APP 商品券或點數支付是否可評估。
4. 場站試點的權責與現勘流程。
"""

FILES[TMP / "VIVI_BD_Positioning.md"] = """# ViVi PARK 專案定位分析

## 定位判斷

ViVi PARK 不是傳統停車場品牌，而是停車場、APP、Vi 幣、加油、EV 充電、點數漫遊與生活服務的整合平台。

旺來切入時不應只說「放置設備」，而應說：

```text
旺來希望提供一個可先小規模驗證的生活服務節點，讓 ViVi PARK 的場站與 Vi 幣生態增加新的日常使用情境。
```

## 為什麼找 ViVi PARK

1. ViVi PARK 已公開主張生活服務與點數生態。
2. 官網明確歡迎生活商家或同業加入。
3. 場站與 APP 會員能形成線上線下閉環。
4. 桃園／台中已有可公開辨識的生活型場站。

## 對 ViVi PARK 的合作價值

| 價值 | 說明 |
|---|---|
| 新使用場景 | 增加 Vi 幣或 APP 會員在停車之外的使用理由 |
| 場站服務力 | 讓停車場從停車點延伸為生活服務觸點 |
| 聯盟案例 | 形成「生活服務接入 ViVi 共榮生態圈」的可示範案例 |
| 數據驗證 | 90 天看使用量、客訴、動線、補貨與會員轉換 |

## 1 至 3 個試點假設

- 城市：桃園／台中。
- 型態：室外平面、住宅／商圈／全聯生活型場站。
- 期間：90 天。
- 決策：雙方共同決定擴大、調整或停止。
"""

FILES[ONE / "VIVI_OnePager.md"] = """# ViVi PARK OnePager 內容稿

> 對象：ViVi PARK／榮帝科技生活服務、異業合作、停車場營運窗口  
> 目標：取得 30 分鐘會議

## 主標

ViVi PARK 已經把停車變成生活入口，旺來希望一起驗證下一個日常服務場景。

## 核心提案

旺來瓦斯希望與 ViVi PARK 交流一個低干擾、可量化、可先小規模驗證的合作方向：

```text
在 1 至 3 個桃園或台中場站，建立「停車後生活服務節點」。
```

本案不是承租停車格，也不是廣告進駐，而是在不影響停車收益、車位與車道動線的前提下，評估是否能將旺來的生活服務能力接入 ViVi PARK 的場站與會員使用場景。

## 為什麼是 ViVi PARK

| 原因 | 說明 |
|---|---|
| 平台定位明確 | ViVi PARK 已整合停車、Vi 幣、加油、EV 充電與生活服務 |
| 生活服務入口 | 官網已開放鄰近店家以商品券或點數支付形式加入 |
| 數位基礎成熟 | APP、Vi 幣、點數漫遊與車辨場站可支撐數據化驗證 |
| 場站可試點 | 桃園／台中有公開生活型候選場站，可先小規模評估 |

## ViVi PARK 可以得到什麼

1. 新增一個可對外說明的生活服務合作案例。
2. 擴大 ViVi PARK APP／Vi 幣的日常使用場景。
3. 用 90 天數據確認場站服務是否真的有需求。
4. 低資本、低干擾、可共同決定是否擴大。

## 旺來負責什麼

設備、客服、補貨、維修、法規符合性、責任保險、異常處理與試點數據整理。

## ViVi PARK 需要協助什麼

提供合作窗口、初步場站條件、權責確認與一次現勘安排。

## 建議下一步

安排 30 分鐘會議，先確認：

1. 合作應走生活服務／特約商店，還是場站試點。
2. 桃園／台中哪些場站可先評估。
3. 90 天試點的資料與責任分工。
"""

FILES[ONE / "VIVI_OnePager_DM.html"] = """<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8" />
<title>ViVi PARK × 旺來瓦斯 OnePager</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#d9d9d9;font-family:"Microsoft JhengHei",Arial,sans-serif;color:#171717}.page{width:794px;height:1123px;margin:0 auto;background:#fff;position:relative;overflow:hidden;padding:54px 58px}.brand{position:absolute;right:50px;top:38px;color:#f05a24;font-weight:800;font-size:28px}.bar{height:7px;background:#f05a24;width:92px;margin-bottom:22px}.eyebrow{color:#f05a24;font-weight:700;font-size:18px}.hero h1{font-size:40px;line-height:1.18;margin:12px 0 18px;letter-spacing:0}.hero p{font-size:18px;line-height:1.65;margin:0;width:520px}.circle{position:absolute;right:-90px;top:120px;width:300px;height:300px;background:#fff1e8;border:1px solid #ffd0ba;border-radius:50%}.node{position:absolute;text-align:center;border:2px solid #f05a24;background:#fff;padding:12px 10px;width:116px}.n1{right:112px;top:178px}.n2{right:44px;top:292px}.n3{right:178px;top:292px}.node b{display:block;color:#f05a24;font-size:25px}.section{margin-top:34px}.section h2{font-size:24px;border-bottom:2px solid #ddd;padding-bottom:10px;margin:0 0 18px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}.card{border:1px solid #ddd;border-left:5px solid #f05a24;padding:16px;min-height:126px}.card h3{margin:0 0 8px;font-size:20px}.card p{margin:0;font-size:15.5px;line-height:1.55}.flow{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}.step{background:#fff7f2;border:1px solid #ffc8ad;padding:15px;min-height:110px}.step b{color:#f05a24;font-size:22px}.step p{font-size:15px;line-height:1.5;margin:8px 0 0}.cta{position:absolute;left:58px;right:58px;bottom:72px;background:#f05a24;color:#fff;padding:22px 26px}.cta h2{margin:0 0 8px;font-size:25px}.cta p{margin:0;font-size:16px;line-height:1.55}.footer{position:absolute;left:58px;right:58px;bottom:28px;border-top:1px solid #ccc;padding-top:10px;font-size:12px;color:#777;display:flex;justify-content:space-between}
</style>
</head>
<body><main class="page">
<div class="brand">waGas</div><div class="bar"></div>
<div class="circle"></div>
<div class="node n1"><b>Vi 幣</b><span>會員場景</span></div>
<div class="node n2"><b>場站</b><span>線下觸點</span></div>
<div class="node n3"><b>旺來</b><span>生活服務</span></div>
<section class="hero">
<div class="eyebrow">ViVi PARK × 旺來瓦斯｜停車後生活服務節點試點合作</div>
<h1>ViVi PARK 已經把停車變成生活入口，旺來希望一起驗證下一個日常服務場景。</h1>
<p>我們希望先從桃園或台中 1 至 3 個示範場站，評估在不影響停車收益、車位與動線的前提下，建立可被 90 天數據驗證的生活服務節點。</p>
</section>
<section class="section"><h2>為什麼這件事適合 ViVi PARK</h2><div class="grid">
<div class="card"><h3>平台定位明確</h3><p>ViVi PARK 已整合停車、Vi 幣、加油、充電與生活服務，合作不只是場地，而是生態圈新增場景。</p></div>
<div class="card"><h3>會員使用延伸</h3><p>旺來能提供日常剛性需求，讓 APP 與 Vi 幣在停車之外有新的使用理由。</p></div>
<div class="card"><h3>場站低干擾</h3><p>先看入口旁、邊角或低使用空間，不佔主要車位，不影響車道動線。</p></div>
<div class="card"><h3>數據可決策</h3><p>90 天後共同看使用量、客訴、補貨、動線與會員反應，再決定是否擴大。</p></div>
</div></section>
<section class="section"><h2>建議試點方式</h2><div class="flow">
<div class="step"><b>1-3 場</b><p>桃園／台中生活型場站優先，先評估條件。</p></div>
<div class="step"><b>90 天</b><p>驗證使用量、異常率、補貨與場站回饋。</p></div>
<div class="step"><b>共同決定</b><p>依結果討論擴大、調整或停止。</p></div>
</div></section>
<section class="cta"><h2>建議先安排 30 分鐘交流</h2><p>確認本案應走生活服務合作、Vi 幣／商品券合作，或場站試點合作；再由雙方共同選擇最合適的 1 至 3 個示範點。</p></section>
<div class="footer"><span>旺來瓦斯股份有限公司｜開發部</span><span>OnePager / VIVI v1</span></div>
</main></body></html>
"""

FILES[COLD / "VIVI_Coldmail_Email.txt"] = """主旨：ViVi PARK × 旺來瓦斯｜停車後生活服務節點試點合作交流

ViVi PARK／榮帝科技團隊 您好：

我是旺來瓦斯股份有限公司開發部門同仁。

近期我們在評估停車場場域與生活服務結合的可能性。ViVi PARK 已從停車服務延伸到 Vi 幣、點數漫遊、加油、EV 充電與生活服務，是我們認為最適合交流「場站服務延伸」的合作對象之一，因此冒昧來信。

本案不是承租停車格，也不是廣告進駐或傳統設備投放。我們希望評估是否能在不影響停車收益、主要車位與車道動線的前提下，於 1 至 3 個桃園或台中場站先建立「停車後生活服務節點」，用 90 天實際數據驗證使用需求與營運可行性。

對 ViVi PARK 而言，這個合作可先聚焦在三個價值：

1. 新增生活服務場景：讓 ViVi PARK APP／Vi 幣在停車之外，多一個日常使用理由。
2. 場站服務升級：讓場站從停車點延伸為社區生活服務觸點。
3. 數據化試點：由雙方依使用量、客訴、補貨、動線與場站回饋，共同決定是否擴大、調整或停止。

旺來可負責設備、客服、補貨、維修、法規符合性、責任保險與異常處理；ViVi PARK 端僅需協助確認合作窗口、場站基本條件、權責與現勘安排。

隨信附上一頁式合作構想，供貴團隊初步參考。若方向合適，是否有機會安排約 30 分鐘線上或實體會議，由我們向生活服務、異業合作或場站營運相關窗口進一步說明？

非常感謝您撥冗閱讀，期待有機會與 ViVi PARK 團隊交流。

敬祝 商祺
旺來瓦斯股份有限公司
開發部
"""

FILES[COLD / "VIVI_Coldmail_Workflow.md"] = """# ViVi PARK Coldmail 工作流

## 目標

取得 ViVi PARK／榮帝科技生活服務、異業合作或場站營運窗口，安排 30 分鐘會議。

## 附件

- `ViVi_PARK_提案專案/01_OnePager/VIVI_OnePager_DM.html`

## 主旨選項

1. ViVi PARK × 旺來瓦斯｜停車後生活服務節點試點合作交流
2. 關於 ViVi PARK 場站與生活服務生態圈合作構想
3. ViVi PARK × 旺來瓦斯｜桃園／台中 1–3 場示範點合作評估

## 寄送前檢查

- 是否明確說明不是租格位、不是廣告、不是設備投放。
- 是否使用 ViVi PARK 的平台語言：Vi 幣、生活服務、共榮生態圈。
- 是否避免一開始要求大量點位。
- 是否收斂到 30 分鐘會議。
"""

FILES[COLD / "VIVI_Coldmail_Package.md"] = """# ViVi PARK Coldmail Package

## 推薦主旨

ViVi PARK × 旺來瓦斯｜停車後生活服務節點試點合作交流

## 正式信件

請使用 `VIVI_Coldmail_Email.txt`。

## 表單短版

您好，我是旺來瓦斯股份有限公司開發部門同仁。

我們希望與 ViVi PARK／榮帝科技交流一項「停車後生活服務節點」試點合作。本案不是承租停車格，也不是廣告進駐，而是在不影響停車收益、車位與動線的前提下，評估 1 至 3 個桃園或台中場站進行小規模驗證。

旺來可負責設備、客服、補貨、維修、法規符合性、責任保險與異常處理；貴司僅需協助確認合作窗口、場站條件與現勘安排。若方向合適，懇請協助轉介生活服務、異業合作或場站營運相關窗口，安排約 30 分鐘會議交流。
"""

FILES[COLD / "VIVI_BD_Framework.md"] = """# ViVi PARK BD Framework

## 主要切入

生活服務生態圈合作，而非單純場地合作。

## 四層價值

1. Vi 幣使用場景。
2. 場站服務延伸。
3. 會員日常需求。
4. 90 天數據驗證。

## 會議中要確認

- 生活服務合作入口。
- 商品券／點數支付可能性。
- 桃園／台中候選場站。
- 場站設置權責與現勘流程。
"""

FILES[COLD / "VIVI_BD_Blueprint.md"] = """# ViVi PARK BD Blueprint

## 路徑

Cold mail → 表單／電話追蹤 → 30 分鐘會議 → 場站條件確認 → 1–3 場現勘 → 90 天試點。

## 反對意見回應

| 反對意見 | 回應 |
|---|---|
| 這是設備進駐嗎？ | 不是，本案是生活服務節點試點，設備只是服務載體。 |
| 會影響場站營運嗎？ | 首波僅評估不影響收益、車位與動線的低使用空間。 |
| 我們已有生活服務合作 | 旺來可作為日常剛性需求場景，補強 Vi 幣／APP 使用理由。 |
"""

FILES[MEET / "VIVI_Followup_Email.txt"] = """主旨：追蹤 ViVi PARK × 旺來瓦斯合作構想資料

ViVi PARK／榮帝科技團隊 您好：

我是旺來瓦斯開發部門同仁。日前有寄送一份「停車後生活服務節點」合作構想，想向貴團隊確認是否已收到資料。

本案主要希望交流 ViVi PARK 場站、APP／Vi 幣與生活服務場景是否有 1 至 3 個示範點可先小規模驗證。若方向合適，想請教是否能協助轉介生活服務、異業合作或場站營運相關窗口，安排約 30 分鐘會議。

謝謝您。
"""

FILES[MEET / "VIVI_Call_Talktrack.md"] = """# ViVi PARK 電話追蹤話術

## 開場

您好，我是旺來瓦斯開發部門同仁，想確認我們寄給 ViVi PARK 的合作構想是否有收到。

## 一句話說明

我們不是要租車位或做廣告，而是想交流一個「停車後生活服務節點」的小規模試點，讓 ViVi PARK 的場站與生活服務生態圈多一個日常使用場景。

## 目標

請窗口協助轉接生活服務、異業合作或場站營運相關同仁，安排 30 分鐘會議。
"""

FILES[MEET / "VIVI_Meeting_Request_Script.md"] = """# ViVi PARK 會議邀約稿

我們希望會議先確認三件事：

1. 本案適合走生活服務／特約商店合作，還是場站試點合作。
2. 桃園或台中是否有 1 至 3 個場站可初步評估。
3. 若要驗證 90 天，雙方各自需要確認哪些資料與權責。
"""

FILES[SHORT / "VIVI_ShortPPT_Outline.md"] = """# ViVi PARK 簡版 PPT 大綱（6–8 頁）

## P1｜ViVi PARK 已把停車變成生活入口
- 核心訊息：ViVi PARK 的價值不只在停車，而在 APP、Vi 幣與生活服務生態圈。
- 視覺：平台生態圈圖。
- 來源：官網首頁、關於我們。

## P2｜旺來可補上一個高頻日常服務場景
- 核心訊息：旺來提供可落地、可補貨、可維運的生活服務能力。
- 視覺：停車後服務動線。

## P3｜合作不是租格位，而是生活服務節點試點
- 核心訊息：不影響停車收益、車位與動線。
- 視覺：不是／而是對照圖。

## P4｜ViVi PARK 可以得到三個價值
- 核心訊息：Vi 幣場景、場站服務升級、數據化案例。
- 視覺：三欄價值矩陣。

## P5｜先用 1 至 3 場、90 天驗證
- 核心訊息：降低決策門檻，先用數據說話。
- 視覺：90 天流程圖。

## P6｜首波候選場站方向
- 核心訊息：桃園／台中生活型、室外平面、全聯或住宅商圈場站優先。
- 視覺：候選條件表。

## P7｜會議希望共同確認
- 核心訊息：合作入口、場站條件、現勘流程。
- 視覺：三項確認清單。
"""

FILES[FORMAL / "VIVI_FormalPPT_Outline.md"] = """# ViVi PARK 正式外部提案大綱（12 頁以內）

## P1｜ViVi PARK × 旺來瓦斯：停車後生活服務節點試點合作
核心訊息：把場站與 Vi 幣生態延伸到新的日常服務場景。

## P2｜停車場競爭正在從停車效率走向生活服務入口
核心訊息：停車、加油、充電、點數、生活服務正在被整合。

## P3｜為什麼是 ViVi PARK
核心訊息：ViVi PARK 已具備平台、會員、點數與生活服務基礎。

## P4｜旺來提供的是可落地的服務營運能力
核心訊息：旺來負責設備、客服、補貨、維修、法遵與保險。

## P5｜合作定位：不是租格位，是場站服務與 Vi 幣場景延伸
核心訊息：避免誤解，建立正確合作框架。

## P6｜ViVi PARK 可以獲得什麼
核心訊息：新增使用場景、場站升級、數據案例。

## P7｜建議從 1 至 3 個桃園／台中場站試點
核心訊息：先小規模驗證，不預設大量展點。

## P8｜90 天驗證方式
核心訊息：用使用量、客訴、補貨、動線與會員反應共同決策。

## P9｜首波候選場站條件
核心訊息：室外平面、生活商圈、全聯／住宅型、可使用 APP 的場站優先。

## P10｜試點前共同確認項目
核心訊息：權責、供電、動線、保險、APP／Vi 幣合作形式。

## P11｜會後建議決策方向
核心訊息：先決定是否安排場站條件確認與窗口對接。
"""

FILES[TMP / "VIVI_Project_Readiness_Check.md"] = """# ViVi PARK Project Readiness Check

## 已完成

- 00_來源資料：DD、BD strategy、Project Master、Headquarter Approach。
- 01_OnePager：內容稿與 HTML DM。
- 02_Coldmail：Email、Package、Workflow、Framework、Blueprint。
- 03_取得窗口與會議：追蹤信、電話話術、會議邀約稿。
- 04/05 PPT：簡版與正式大綱。

## 待確認

1. ViVi PARK 官方完整桃園／台中場站清單。
2. 生活服務合作是否可走商品券或 Vi 幣支付。
3. 場站試點與 APP 合作是否屬同一窗口。
4. 旺來是否要先寄客服信箱，或同步走官網／LINE@。
"""

FILES[TMP / "VIVI_BD_AUTOMATION_PROMPTS.md"] = """# ViVi PARK BD 自動化工作流 Prompt

節點：

- -1 DD 調查與來源資料
- -0.5 案場資料庫
- 0 專案盤點
- 1 定位分析
- 2 OnePager
- 3 OnePager HTML
- 5 Coldmail
- 6 追蹤話術
- 7 簡版 PPT 大綱
- 8 正式 PPT 大綱

每節完成後應停下等待確認。若使用者要求強烈執行，可連續產出，但每個節點需回報狀態。
"""


def write_files():
    for path, content in FILES.items():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")


def style(ws, widths):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def make_excel():
    wb = Workbook()
    sheets = ["01_Company_DD", "02_Decision_Map", "03_台中場站庫", "04_桃園場站庫", "05_評分模型", "06_TOP10初選", "07_供電分析", "08_Risk_Register", "09_Action_Plan", "10_Source_Assumptions"]
    wb.active.title = sheets[0]
    for name in sheets[1:]:
        wb.create_sheet(name)

    ws = wb["01_Company_DD"]
    rows = [
        ["項目", "內容", "BD含意", "資料來源"],
        ["品牌", "ViVi PARK", "智慧停車與生活服務平台", "官網"],
        ["法人", "榮帝科技股份有限公司", "以總部與平台合作切入", "公司登記／官網"],
        ["統編", "24963309", "正式開發對象", "公司登記"],
        ["董事長", "陳明賢 Paul Chen", "品牌敘事集中於平台願景", "官網關於我們"],
        ["聯絡", "02-2536-6655 / service-vivipark@kpclc.com.tw", "可作 cold mail 與追蹤入口", "官網頁尾"],
        ["核心能力", "Vi 幣、APP、點數漫遊、EV、加油、生活服務", "旺來可接入生活服務場景", "官網"],
    ]
    for r in rows: ws.append(r)
    style(ws, [16, 42, 48, 28])

    ws = wb["02_Decision_Map"]
    for r in [
        ["層級", "角色", "可能關注", "旺來準備", "接觸方式", "狀態"],
        ["L1", "董事長／高階主管", "生態圈策略、品牌方向", "外部提案簡版", "不作第一線", "後期"],
        ["L2", "生活服務／聯盟合作主管", "商品券、點數、會員導流", "OnePager＋合作模型", "主要切入", "待聯繫"],
        ["L3", "場站營運／開發主管", "權責、動線、供電", "案場資料庫＋現勘表", "會議後導入", "待確認"],
        ["L4", "客服／行政窗口", "轉介正確部門", "短版說明", "客服信箱／電話", "待執行"],
    ]: ws.append(r)
    style(ws, [10, 24, 34, 34, 22, 14])

    tc = [
        ["VIVI-TC-001", "全聯台中潭子福潭店", "台中市潭子區福潭路362、368號", "潭子區", "室外平面／全聯場", "是", "中【權責待確認】", "是", "是", "否", "待驗證", "高【車辨設備】", 5, 2, "B+", "官方新場資訊；全聯生活場域，權責需確認"],
        ["VIVI-TC-002", "全聯台中大里爽文店", "台中市大里區爽文路879號", "大里區", "室外平面／全聯場", "是", "中【權責待確認】", "是", "是", "否", "待驗證", "高【車辨設備】", 5, 2, "B+", "官方新場資訊；生活型場站"],
        ["VIVI-TC-003", "台中市區生活型場站_待官方清單", "待 ViVi PARK 提供", "台中市", "待驗證", "待現勘", "待現勘", "待驗證", "待驗證", "待驗證", "待驗證", "待驗證", 3, 3, "B", "需由官網場站查詢或總部提供完整清單"],
        ["VIVI-TC-004", "台中商圈／住宅場站_待官方清單", "待 ViVi PARK 提供", "台中市", "待驗證", "待現勘", "待現勘", "是", "是", "待驗證", "待驗證", "待驗證", 3, 3, "B", "補充備選"],
        ["VIVI-TC-005", "台中全聯／零售場站_待官方清單", "待 ViVi PARK 提供", "台中市", "零售場域", "待現勘", "權責待確認", "是", "是", "否", "待驗證", "中", 4, 3, "B", "若為全聯合作場，需確認第三方權責"],
    ]
    ty = [
        ["VIVI-TY-001", "同安街場", "桃園市桃園區同安街322號", "桃園區", "臨停／月租，APP 可用", "待現勘", "中高【待現勘】", "是", "是", "否", "待驗證", "中高【APP 可用】", 5, 2, "B+", "官方新場資訊；鄰近藝文特區與圖書館"],
        ["VIVI-TY-002", "全聯平鎮中豐店", "桃園市平鎮區中豐路南勢二段326號", "平鎮區", "室外平面／全聯場", "是", "中【權責待確認】", "是", "是", "否", "待驗證", "高【車辨／APP】", 5, 2, "B+", "官方新場資訊；生活型零售場域"],
        ["VIVI-TY-003", "中壢松義二街停車場", "桃園市中壢區松義二街32號旁", "中壢區", "汽車40格，臨停／月租", "是", "高【旁空地】", "是", "是", "否", "待驗證", "中高【APP 可用】", 5, 1, "A", "官方新場資訊；旁空地且住宅／活動場域，首波優先"],
        ["VIVI-TY-004", "桃園水岸五街二停車場", "地址待官方確認", "桃園區", "月租公開資訊", "待現勘", "待現勘", "是", "待驗證", "否", "待驗證", "待驗證", 3, 3, "B", "第三方月租平台出現，需官方確認"],
        ["VIVI-TY-005", "桃園愛買停車場", "地址待官方確認", "桃園區", "零售／商場場站", "待現勘", "權責待確認", "是", "是", "否", "待驗證", "中", 4, 3, "B", "第三方月租平台出現，需官方確認"],
    ]
    headers = ["編號", "場站名稱", "地址", "區域", "場站型態", "是否開放空間", "邊角空間", "周邊住宅", "商圈", "市場", "學區", "供電訊號", "合作價值", "導入難度", "優先級", "備註"]
    for sheet, data in [("03_台中場站庫", tc), ("04_桃園場站庫", ty)]:
        ws = wb[sheet]
        ws.append(headers)
        for r in data: ws.append(r)
        style(ws, [14, 24, 34, 14, 20, 14, 18, 12, 10, 10, 10, 22, 10, 10, 10, 48])

    ws = wb["05_評分模型"]
    for r in [
        ["模型項目", "規則", "本案操作方式"],
        ["合作價值 5", "住宅／零售／商圈＋APP 或平面場", "優先列入首波"],
        ["合作價值 4", "生活圈明確但權責待確認", "列備選"],
        ["導入難度 1", "旁空地／平面且低干擾", "可快速現勘"],
        ["導入難度 2", "平面但權責需確認", "B+ 候選"],
        ["導入難度 3", "地址或型態待補", "B 候選"],
        ["導入難度 5", "地下／商場權責複雜", "暫不首波"],
    ]: ws.append(r)
    style(ws, [18, 34, 58])

    ws = wb["06_TOP10初選"]
    for r in [
        ["排序", "編號", "城市", "場站名稱", "推薦原因", "主要風險", "現勘必查", "提案語"],
        [1, "VIVI-TY-003", "桃園", "中壢松義二街停車場", "旁空地、40格、臨停／月租，生活圈清楚", "供電與權責需確認", "電箱／邊角／補貨短停", "首波 A 級候選"],
        [2, "VIVI-TY-001", "桃園", "同安街場", "桃園藝文特區生活場景，APP 可用", "動線與邊角待現勘", "動線／電箱／設置位置", "桃園核心候選"],
        [3, "VIVI-TY-002", "桃園", "全聯平鎮中豐店", "全聯生活場域，會員與消費場景契合", "全聯／ViVi PARK 權責", "第三方同意／用電", "零售型候選"],
        [4, "VIVI-TC-001", "台中", "全聯台中潭子福潭店", "室外平面、全聯場、車辨設備", "權責與設置規範", "供電／權責／動線", "台中 B+ 候選"],
        [5, "VIVI-TC-002", "台中", "全聯台中大里爽文店", "生活型零售場域，APP 可用", "權責與補貨", "電箱／權責／補貨路線", "台中 B+ 候選"],
    ]: ws.append(r)
    style(ws, [8, 14, 10, 24, 44, 34, 34, 38])

    ws = wb["07_供電分析"]
    for r in [
        ["編號", "場站名稱", "Availability 有沒有電", "Accessibility 能不能接", "Economics 是否划算", "分數", "判斷", "下一步"],
        ["VIVI-TY-003", "中壢松義二街", "APP 可用、臨停／月租，推測有基本設備電源", "旁空地可能較易安排，但需電箱位置", "短距離可行；新增電表降級", "4", "優先現勘", "請 ViVi PARK 提供電箱位置"],
        ["VIVI-TY-001", "同安街場", "APP 可用，推測有設備電源", "需避開藝文區場站動線", "可控", "3", "可評估", "現勘確認邊角"],
        ["VIVI-TY-002", "全聯平鎮中豐", "車辨／APP，供電可能充足", "全聯權責需確認", "審核成本可能高", "3", "權責優先", "先確認合作主體"],
        ["VIVI-TC-001", "全聯潭子福潭", "車辨設備，供電可能充足", "全聯權責需確認", "審核成本可能高", "3", "權責優先", "先確認合作主體"],
    ]: ws.append(r)
    style(ws, [14, 22, 34, 34, 34, 10, 28, 34])

    ws = wb["08_Risk_Register"]
    for r in [
        ["風險", "描述", "機率", "衝擊", "等級", "緩解措施", "Owner", "狀態"],
        ["平台合作權責", "生活服務、點數、場站可能分屬不同窗口", "高", "中", "黃", "先釐清合作入口", "BD", "待確認"],
        ["第三方場站權責", "全聯或商場場站需第三方同意", "高", "高", "紅", "首波先確認權責", "BD/法務", "待確認"],
        ["被誤認為設備投放", "對方可能以場地設備審核處理", "中", "中", "黃", "明確說生活服務節點", "BD", "可管控"],
        ["Vi 幣整合門檻", "點數或商品券清算需規則", "中", "中", "黃", "先不承諾支付整合", "BD/財務", "待確認"],
    ]: ws.append(r)
    style(ws, [20, 42, 10, 10, 10, 38, 14, 14])

    ws = wb["09_Action_Plan"]
    for r in [
        ["階段", "時間", "任務", "負責人", "輸入", "輸出", "成功標準", "狀態"],
        ["Phase 1", "D+1", "寄送 Coldmail + OnePager", "BD", "OnePager", "寄件紀錄", "收到或轉介", "待辦"],
        ["Phase 1", "D+2", "電話追蹤客服信箱", "BD", "寄件紀錄", "窗口", "取得轉介", "待辦"],
        ["Phase 2", "D+5", "30 分鐘會議", "BD", "簡報大綱", "會議紀錄", "確認合作入口", "待辦"],
        ["Phase 3", "D+7", "場站資料與現勘", "BD/工程", "案場資料庫", "現勘清單", "選出 1–3 點", "待辦"],
    ]: ws.append(r)
    style(ws, [14, 12, 32, 14, 24, 24, 28, 14])

    ws = wb["10_Source_Assumptions"]
    for r in [
        ["來源", "用途", "可信度", "備註"],
        ["ViVi PARK 官網", "品牌、聯絡、平台定位", "高", "官方來源"],
        ["ViVi PARK 關於我們", "公司、董事長、規模、里程碑", "高", "數字需會議確認最新版口徑"],
        ["ViVi PARK 新場資訊", "桃園／台中候選場站", "高", "官方新聞頁"],
        ["第三方月租平台", "補充候選站名", "中", "需官方確認"],
    ]: ws.append(r)
    style(ws, [40, 34, 12, 50])

    wb.save(SRC / "VIVI_案場資料庫_v1.xlsx")


def main():
    write_files()
    make_excel()
    print("VIVI project generated")


if __name__ == "__main__":
    main()
